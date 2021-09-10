import openpyxl
from openpyxl.styles import Alignment
import os
import pathlib
import re
import pandas as pd
import time


def copy_range(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow, endRow + 1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        rangeSelected.append(rowSelected)

    return rangeSelected


def paste_range(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


def main():
    xlsx_list = []
    xlsx_list_cleaned = []

    # data_to_be_processed = "data_to_be_processed"
    cleaned_data_dir = "cleaned_data"
    combined_data_dir = "combined_data"

    # if not os.path.exists(data_to_be_processed):
    #     os.makedirs(data_to_be_processed)

    if not os.path.exists(cleaned_data_dir):
        os.makedirs(cleaned_data_dir)

    if not os.path.exists(combined_data_dir):
        os.makedirs(combined_data_dir)

    time.sleep(1)

    # os.chdir("data_to_be_processed")

    xlsx = pathlib.Path().glob("*.xlsx")
    xlsx_cleaned = pathlib.Path("cleaned_data").glob("*.xlsx")

    for file in xlsx:
        xlsx_list.append(file)
        print(file)

    for i in xlsx_list:
        file = f"{i}"
        cleaned_file = f"{file} Cleaned.xlsx"

        if os.path.exists(cleaned_file):
            os.remove(cleaned_file)

        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb['Result Sheet']

        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        sheet2 = wb2['Sheet']
        sheet2.title = 'Results Sheet'

        # for rowOfCellObjects in sheet['A25': 'A51']:
        #     for cellObj in rowOfCellObjects:
        #         registration_num_data.append(cellObj.value)
        #
        # print(registration_num_data)

        registration_range = copy_range(1, 25, 1, 51, sheet)
        paste_range(1, 2, 1, 26, sheet2, registration_range)

        name_of_students_range = copy_range(2, 25, 2, 51, sheet)
        paste_range(2, 2, 2, 26, sheet2, name_of_students_range)

        total_marks_range = copy_range(8, 25, 8, 51, sheet)
        paste_range(3, 2, 3, 26, sheet2, total_marks_range)

        grades_range = copy_range(9, 25, 9, 51, sheet)
        paste_range(4, 2, 4, 26, sheet2, grades_range)

        for row in ws2.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        sheet2['A1'] = "registration_num"
        sheet2['B1'] = "student_name"
        sheet2['C1'] = "total_num"
        sheet2['D1'] = "grade"
        sheet2['E1'] = "year"

        pattern = re.search("^[0-9]{4}", file)
        # print(pattern.group())

        for j in range(2, 27):
            ws2[f"E{j}"].value = int(pattern.group())

        os.chdir('cleaned_data')
        wb2.save(filename=f'{cleaned_file}')
        os.chdir('../')

    for file in xlsx_cleaned:
        xlsx_list_cleaned.append(file)
        print(file)

    excels = [pd.ExcelFile(name) for name in xlsx_list_cleaned]
    frames = [x.parse(x.sheet_names[0], header=None, index_col=None) for x in excels]
    frames[1:] = [df[1:] for df in frames[1:]]
    combined = pd.concat(frames)

    os.chdir('combined_data')
    combined.to_excel("combined.xlsx", header=False, index=False)
    os.chdir('../')


if __name__ == '__main__':
    main()
